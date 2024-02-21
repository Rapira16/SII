import openpyxl
import matplotlib.pyplot as plt

book = openpyxl.load_workbook(filename="C:/Users/DIMOOON/OneDrive/Рабочий стол/ВУЗ/apple.xlsx")
sheet_obj = book.active
m_row = sheet_obj.max_row
date = []
apple = []
microsoft = []
google = []

for i in range(2, m_row+1):
    cell_obj_date = sheet_obj.cell(row = i, column = 1)
    cell_obj_apple = sheet_obj.cell(row = i, column = 2)
    cell_obj_google = sheet_obj.cell(row=i, column=3)
    cell_obj_microsoft = sheet_obj.cell(row=i, column=4)
    date.append(cell_obj_date.value)
    apple.append(float(cell_obj_apple.value))
    google.append(float(cell_obj_google.value))
    microsoft.append(float(cell_obj_microsoft.value))

data = {'apple':apple, 'google':google, 'microsoft':microsoft}
plt.plot(date, apple)
plt.plot(date, google)
plt.plot(date, microsoft)
plt.legend(data, loc=2)
plt.show()